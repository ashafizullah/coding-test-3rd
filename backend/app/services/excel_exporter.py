"""
Excel Export Service

Generates Excel files with fund data, transactions, and metrics.
"""
from typing import Optional
from datetime import datetime
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.fund import Fund
from app.models.transaction import CapitalCall, Distribution, Adjustment
from app.services.metrics_calculator import MetricsCalculator

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export fund data to Excel format"""

    def __init__(self, db: Session):
        self.db = db
        self.metrics_calculator = MetricsCalculator(db)

    def export_fund_report(self, fund_id: int) -> bytes:
        """
        Generate comprehensive fund report in Excel format.

        Args:
            fund_id: Fund ID to export

        Returns:
            Excel file as bytes
        """
        # Get fund data
        fund = self.db.query(Fund).filter(Fund.id == fund_id).first()
        if not fund:
            raise ValueError(f"Fund with id {fund_id} not found")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add sheets
        self._create_summary_sheet(wb, fund)
        self._create_metrics_sheet(wb, fund)
        self._create_capital_calls_sheet(wb, fund)
        self._create_distributions_sheet(wb, fund)
        self._create_adjustments_sheet(wb, fund)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Generated Excel report for fund {fund_id}")
        return buffer.getvalue()

    def _create_summary_sheet(self, wb: Workbook, fund: Fund):
        """Create summary sheet with fund information"""
        ws = wb.create_sheet("Summary")

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)

        # Title
        ws["A1"] = fund.name
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:D1")

        # Fund Information
        ws["A3"] = "Fund Information"
        ws["A3"].font = header_font
        ws["A3"].fill = header_fill
        ws.merge_cells("A3:B3")

        info_rows = [
            ("GP Name:", fund.gp_name or "N/A"),
            ("Fund Type:", fund.fund_type or "N/A"),
            ("Vintage Year:", str(fund.vintage_year) if fund.vintage_year else "N/A"),
            ("Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]

        row = 4
        for label, value in info_rows:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Key Metrics
        ws[f"A{row + 1}"] = "Key Metrics"
        ws[f"A{row + 1}"].font = header_font
        ws[f"A{row + 1}"].fill = header_fill
        ws.merge_cells(f"A{row + 1}:B{row + 1}")

        try:
            metrics = self.metrics_calculator.calculate_all_metrics(fund.id)

            metrics_rows = [
                ("DPI:", f"{metrics.get('dpi', 0):.2f}x" if metrics.get('dpi') else "N/A"),
                ("IRR:", f"{metrics.get('irr', 0) * 100:.2f}%" if metrics.get('irr') else "N/A"),
                ("Paid-In Capital:", f"${metrics.get('pic', 0):,.2f}" if metrics.get('pic') else "N/A"),
                ("Total Distributions:", f"${metrics.get('total_distributions', 0):,.2f}" if metrics.get('total_distributions') else "N/A"),
            ]

            row += 2
            for label, value in metrics_rows:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = value
                row += 1
        except Exception as e:
            logger.error(f"Error calculating metrics: {e}")

        # Auto-size columns
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _create_metrics_sheet(self, wb: Workbook, fund: Fund):
        """Create detailed metrics breakdown sheet"""
        ws = wb.create_sheet("Metrics Breakdown")

        try:
            breakdown = self.metrics_calculator.get_calculation_breakdown(fund.id)

            # Headers
            ws["A1"] = "Metrics Calculation Breakdown"
            ws["A1"].font = Font(bold=True, size=14)

            row = 3
            # DPI Calculation
            ws[f"A{row}"] = "DPI Calculation"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            dpi_data = breakdown.get("dpi_calculation", {})
            ws[f"A{row}"] = "Total Distributions:"
            ws[f"B{row}"] = f"${dpi_data.get('total_distributions', 0):,.2f}"
            row += 1
            ws[f"A{row}"] = "Paid-In Capital:"
            ws[f"B{row}"] = f"${dpi_data.get('pic', 0):,.2f}"
            row += 1
            ws[f"A{row}"] = "DPI:"
            ws[f"B{row}"] = f"{dpi_data.get('dpi', 0):.4f}x"
            ws[f"B{row}"].font = Font(bold=True)

            row += 3
            # IRR Calculation
            ws[f"A{row}"] = "IRR Calculation"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            irr_data = breakdown.get("irr_calculation", {})
            ws[f"A{row}"] = "IRR:"
            ws[f"B{row}"] = f"{irr_data.get('irr', 0) * 100:.2f}%" if irr_data.get('irr') else "N/A"
            ws[f"B{row}"].font = Font(bold=True)
            row += 1
            ws[f"A{row}"] = "Number of Cash Flows:"
            ws[f"B{row}"] = irr_data.get('cash_flow_count', 0)

        except Exception as e:
            logger.error(f"Error creating metrics breakdown: {e}")
            ws["A3"] = f"Error generating breakdown: {str(e)}"

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_capital_calls_sheet(self, wb: Workbook, fund: Fund):
        """Create capital calls transaction sheet"""
        ws = wb.create_sheet("Capital Calls")

        # Headers
        headers = ["Date", "Type", "Amount", "Description"]
        self._write_table_headers(ws, headers, 1)

        # Get data
        capital_calls = self.db.query(CapitalCall).filter(
            CapitalCall.fund_id == fund.id
        ).order_by(CapitalCall.call_date).all()

        # Write data
        row = 2
        total = 0
        for call in capital_calls:
            ws[f"A{row}"] = call.call_date.strftime("%Y-%m-%d")
            ws[f"B{row}"] = call.call_type or ""
            ws[f"C{row}"] = float(call.amount)
            ws[f"C{row}"].number_format = '"$"#,##0.00'
            ws[f"D{row}"] = call.description or ""
            total += float(call.amount)
            row += 1

        # Total row
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"] = total
        ws[f"C{row}"].number_format = '"$"#,##0.00'
        ws[f"C{row}"].font = Font(bold=True)

        self._auto_size_columns(ws, headers)

    def _create_distributions_sheet(self, wb: Workbook, fund: Fund):
        """Create distributions transaction sheet"""
        ws = wb.create_sheet("Distributions")

        # Headers
        headers = ["Date", "Type", "Amount", "Recallable", "Description"]
        self._write_table_headers(ws, headers, 1)

        # Get data
        distributions = self.db.query(Distribution).filter(
            Distribution.fund_id == fund.id
        ).order_by(Distribution.distribution_date).all()

        # Write data
        row = 2
        total = 0
        for dist in distributions:
            ws[f"A{row}"] = dist.distribution_date.strftime("%Y-%m-%d")
            ws[f"B{row}"] = dist.distribution_type or ""
            ws[f"C{row}"] = float(dist.amount)
            ws[f"C{row}"].number_format = '"$"#,##0.00'
            ws[f"D{row}"] = "Yes" if dist.is_recallable else "No"
            ws[f"E{row}"] = dist.description or ""
            total += float(dist.amount)
            row += 1

        # Total row
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"] = total
        ws[f"C{row}"].number_format = '"$"#,##0.00'
        ws[f"C{row}"].font = Font(bold=True)

        self._auto_size_columns(ws, headers)

    def _create_adjustments_sheet(self, wb: Workbook, fund: Fund):
        """Create adjustments transaction sheet"""
        ws = wb.create_sheet("Adjustments")

        # Headers
        headers = ["Date", "Type", "Category", "Amount", "Description"]
        self._write_table_headers(ws, headers, 1)

        # Get data
        adjustments = self.db.query(Adjustment).filter(
            Adjustment.fund_id == fund.id
        ).order_by(Adjustment.adjustment_date).all()

        # Write data
        row = 2
        total = 0
        for adj in adjustments:
            ws[f"A{row}"] = adj.adjustment_date.strftime("%Y-%m-%d")
            ws[f"B{row}"] = adj.adjustment_type or ""
            ws[f"C{row}"] = adj.category or ""
            ws[f"D{row}"] = float(adj.amount)
            ws[f"D{row}"].number_format = '"$"#,##0.00'
            ws[f"E{row}"] = adj.description or ""
            total += float(adj.amount)
            row += 1

        # Total row
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"D{row}"] = total
        ws[f"D{row}"].number_format = '"$"#,##0.00'
        ws[f"D{row}"].font = Font(bold=True)

        self._auto_size_columns(ws, headers)

    def _write_table_headers(self, ws, headers, row):
        """Write table headers with styling"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_size_columns(self, ws, headers):
        """Auto-size columns based on content"""
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(len(header) + 2, 15)
